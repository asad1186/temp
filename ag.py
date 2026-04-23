import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---- CONFIG ---- #
MAIN_FILE = 'main_file.xlsx'       # Your main GBT file
CONCUR_FILE = 'xyz.xl'             # Your concur file
OWNER_COL_MAIN = 'Owner Name'      # Owner column in main file  <-- change
OWNER_COL_CONCUR = 'Owner Name'    # Owner column in xyz.xl    <-- change if different
OUTPUT_FILE = 'output.xlsx'

Here's the complete updated code with just bold headers (no colors):
pythonimport pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ---- CONFIG ---- #
MAIN_FILE = 'main_file.xlsx'       # Your main GBT file
CONCUR_FILE = 'xyz.xl'             # Your concur file
OWNER_COL_MAIN = 'Owner Name'      # Owner column in main file  <-- change
OWNER_COL_CONCUR = 'Owner Name'    # Owner column in xyz.xl    <-- change if different
OUTPUT_FILE = 'output.xlsx'

# ---- LOAD DATA ---- #
df_main = pd.read_excel(MAIN_FILE, sheet_name=0)
df_concur = pd.read_excel(CONCUR_FILE, sheet_name=0)

wb = load_workbook(MAIN_FILE)
ws = wb.create_sheet('Grouped by Owner')

main_headers = list(df_main.columns)
concur_headers = list(df_concur.columns)

current_row = 1

def write_header_row(ws, row, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

def write_label(ws, row, label, size=13):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, size=size)

# ---- LOOP THROUGH EACH OWNER ---- #
all_owners = df_main[OWNER_COL_MAIN].dropna().unique()

for owner in all_owners:

    # 1. Owner name
    write_label(ws, current_row, owner)
    current_row += 1

    # 2. GBT Records label
    write_label(ws, current_row, 'GBT Records', size=11)
    current_row += 1

    # 3. GBT column headers (bold only)
    write_header_row(ws, current_row, main_headers)
    current_row += 1

    # 4. GBT data rows for this owner
    owner_gbt = df_main[df_main[OWNER_COL_MAIN] == owner]
    for _, row in owner_gbt.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1

    # 5. One blank row
    current_row += 1

    # 6. AB Records label
    write_label(ws, current_row, 'AB Records', size=11)
    current_row += 1

    # 7. One blank row
    current_row += 1

    # 8. Concur Records label
    write_label(ws, current_row, 'Concur Records', size=11)
    current_row += 1

    # 9. Concur column headers (bold only)
    write_header_row(ws, current_row, concur_headers)
    current_row += 1

    # 10. Concur data rows for this owner
    owner_concur = df_concur[df_concur[OWNER_COL_CONCUR] == owner]
    if not owner_concur.empty:
        for _, row in owner_concur.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
    else:
        current_row += 1  # one blank row if no concur data found

    # 11. Two blank rows before next owner
    current_row += 2

# ---- AUTO FIT COLUMNS ---- #
for col in ws.columns:
    max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

wb.save(OUTPUT_FILE)
print("Done! Check the 'Grouped by Owner' sheet in", OUTPUT_FILE)

wb.save(OUTPUT_FILE)
print("Done! Check the 'Grouped by Owner' sheet in", OUTPUT_FILE)
